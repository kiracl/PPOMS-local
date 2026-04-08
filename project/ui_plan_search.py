from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, 
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox, QMessageBox,
    QFileDialog, QAbstractItemView, QFrame
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor
import database
import openpyxl
from datetime import datetime

class PlanSearchWidget(QWidget):
    request_open_detail = Signal(str)

    def __init__(self):
        super().__init__()
        self.page = 1
        self.page_size = 20
        self.total_count = 0
        self.total_pages = 1
        self.sort_by = None
        self.sort_desc = False
        
        self.init_ui()
        self.load_demand_units()
        self.load_data()

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 1. Filter Area
        filter_group = QFrame()
        filter_group.setFrameShape(QFrame.StyledPanel)
        filter_layout = QHBoxLayout(filter_group)
        
        self.seq_input = QLineEdit()
        self.seq_input.setPlaceholderText("序号 (精确)")
        self.seq_input.returnPressed.connect(self.on_search)
        
        self.item_input = QLineEdit()
        self.item_input.setPlaceholderText("采购标的 (模糊)")
        self.item_input.returnPressed.connect(self.on_search)
        
        self.order_input = QLineEdit()
        self.order_input.setPlaceholderText("主单编号 (精确)")
        self.order_input.returnPressed.connect(self.on_search)
        
        self.unit_combo = QComboBox()
        self.unit_combo.addItem("全部")
        
        filter_layout.addWidget(QLabel("序号:"))
        filter_layout.addWidget(self.seq_input)
        filter_layout.addWidget(QLabel("采购标的:"))
        filter_layout.addWidget(self.item_input)
        filter_layout.addWidget(QLabel("主单编号:"))
        filter_layout.addWidget(self.order_input)
        filter_layout.addWidget(QLabel("需求单位:"))
        filter_layout.addWidget(self.unit_combo)
        
        btn_search = QPushButton("查询")
        btn_search.clicked.connect(self.on_search)
        filter_layout.addWidget(btn_search)
        
        btn_reset = QPushButton("重置")
        btn_reset.clicked.connect(self.on_reset)
        filter_layout.addWidget(btn_reset)
        
        layout.addWidget(filter_group)
        
        # 2. Toolbar (Import/Export)
        toolbar = QHBoxLayout()
        
        btn_import = QPushButton("导入Excel")
        btn_import.clicked.connect(self.on_import)
        toolbar.addWidget(btn_import)
        
        btn_export = QPushButton("导出Excel")
        btn_export.clicked.connect(self.on_export)
        toolbar.addWidget(btn_export)
        
        btn_sync = QPushButton("同步系统数据")
        btn_sync.clicked.connect(self.on_sync_system)
        toolbar.addWidget(btn_sync)
        
        toolbar.addStretch()
        layout.addLayout(toolbar)
        
        # 3. Data Table
        self.table = QTableWidget()
        self.columns = ["序号", "主单编号", "需求单位", "采购标的", "规格型号", "采购数量", "单位", "计划发放日期", "计划发放"]
        self.db_fields = ["sequence_no", "main_order_no", "demand_unit", "item_name", "spec_model", "qty", "unit", "plan_date", "plan_release"]
        
        self.table.setColumnCount(len(self.columns))
        self.table.setHorizontalHeaderLabels(self.columns)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.cellDoubleClicked.connect(self.on_cell_double_clicked)
        
        # Sorting
        self.table.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        
        layout.addWidget(self.table)
        
        # 4. Pagination
        page_layout = QHBoxLayout()
        self.btn_prev = QPushButton("上一页")
        self.btn_prev.clicked.connect(self.prev_page)
        self.btn_next = QPushButton("下一页")
        self.btn_next.clicked.connect(self.next_page)
        self.lbl_page = QLabel("第 1 / 1 页 (共 0 条)")
        
        page_layout.addStretch()
        page_layout.addWidget(self.btn_prev)
        page_layout.addWidget(self.lbl_page)
        page_layout.addWidget(self.btn_next)
        page_layout.addStretch()
        
        layout.addLayout(page_layout)

    def on_cell_double_clicked(self, row, column):
        # Column 1 is "主单编号"
        item = self.table.item(row, 1)
        if item:
            number = item.text().strip()
            if number:
                self.request_open_detail.emit(number)
            else:
                QMessageBox.warning(self, "提示", "该记录没有主单编号，无法跳转")

    def load_demand_units(self):
        units = database.fetch_units()
        self.unit_combo.addItems(units)

    def load_data(self):
        seq = self.seq_input.text().strip()
        item = self.item_input.text().strip()
        order = self.order_input.text().strip()
        unit = self.unit_combo.currentText()
        
        # Convert sort column index to field name if sorted
        sort_field = None
        if self.sort_by is not None:
            sort_field = self.db_fields[self.sort_by]
            
        rows, total = database.fetch_plan_search_items(
            filter_seq=seq,
            filter_item=item,
            filter_order=order,
            filter_unit=unit,
            page=self.page,
            page_size=self.page_size,
            sort_by=sort_field,
            sort_desc=self.sort_desc
        )
        
        self.total_count = total
        self.total_pages = (total + self.page_size - 1) // self.page_size
        if self.total_pages < 1: self.total_pages = 1
        
        self.update_pagination_labels()
        self.populate_table(rows)

    def populate_table(self, rows):
        self.table.setRowCount(0)
        for row_idx, row_data in enumerate(rows):
            self.table.insertRow(row_idx)
            for col_idx, val in enumerate(row_data):
                item = QTableWidgetItem(str(val) if val is not None else "")
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, col_idx, item)

    def update_pagination_labels(self):
        self.lbl_page.setText(f"第 {self.page} / {self.total_pages} 页 (共 {self.total_count} 条)")
        self.btn_prev.setEnabled(self.page > 1)
        self.btn_next.setEnabled(self.page < self.total_pages)

    def on_search(self):
        self.page = 1
        self.load_data()

    def on_reset(self):
        self.seq_input.clear()
        self.item_input.clear()
        self.order_input.clear()
        self.unit_combo.setCurrentIndex(0)
        self.page = 1
        self.sort_by = None
        self.sort_desc = False
        self.load_data()

    def prev_page(self):
        if self.page > 1:
            self.page -= 1
            self.load_data()

    def next_page(self):
        if self.page < self.total_pages:
            self.page += 1
            self.load_data()

    def on_header_clicked(self, index):
        if self.sort_by == index:
            self.sort_desc = not self.sort_desc
        else:
            self.sort_by = index
            self.sort_desc = False # Default ASC for new column
        self.load_data()

    def on_import(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel Files (*.xlsx *.xls)")
        if not path:
            return
            
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet = wb.active
            
            # Assuming header is row 1
            headers = [cell.value for cell in sheet[1]]
            required_headers = ["序号", "主单编号", "需求单位", "采购标的", "规格型号", "采购数量", "单位", "计划发放日期", "计划发放"]
            
            # Map headers to col index
            col_map = {}
            for i, h in enumerate(headers):
                if h:
                    col_map[str(h).strip()] = i
            
            # Check missing
            missing = [h for h in required_headers if h not in col_map]
            if missing:
                # If only plan_release is missing, it might be an old template, we can tolerate it or warn
                if missing == ["计划发放"]:
                    pass
                else:
                    QMessageBox.warning(self, "导入失败", f"缺少列: {', '.join(missing)}")
                    return
                
            data_list = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Check sequence (required)
                seq = row[col_map["序号"]]
                if not seq:
                    continue # Skip empty sequence
                
                item = {
                    "sequence_no": str(seq).strip(),
                    "main_order_no": str(row[col_map["主单编号"]] or "").strip(),
                    "demand_unit": str(row[col_map["需求单位"]] or "").strip(),
                    "item_name": str(row[col_map["采购标的"]] or "").strip(),
                    "spec_model": str(row[col_map["规格型号"]] or "").strip(),
                    "qty": row[col_map["采购数量"]],
                    "unit": str(row[col_map["单位"]] or "").strip(),
                    "plan_date": str(row[col_map["计划发放日期"]] or "").strip(),
                    "plan_release": str(row[col_map.get("计划发放")] and row[col_map["计划发放"]] or "").strip() if "计划发放" in col_map else ""
                }
                
                # Basic validation logic could go here
                data_list.append(item)
                
            if not data_list:
                QMessageBox.information(self, "提示", "未找到有效数据")
                return
                
            reply = QMessageBox.question(
                self, 
                "确认导入", 
                f"将导入 {len(data_list)} 条数据。\n重复的序号将覆盖现有记录，新序号将添加。\n是否继续？",
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                inserted, updated = database.import_plan_search_items(data_list)
                QMessageBox.information(self, "导入成功", f"新增: {inserted} 条\n更新: {updated} 条")
                self.on_search()
                
        except Exception as e:
            QMessageBox.critical(self, "导入出错", str(e))

    def on_export(self):
        path, _ = QFileDialog.getSaveFileName(self, "导出Excel", f"计划检索导出_{datetime.now().strftime('%Y%m%d')}.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
            
        try:
            # Get all matching data (not just current page)
            seq = self.seq_input.text().strip()
            item = self.item_input.text().strip()
            order = self.order_input.text().strip()
            unit = self.unit_combo.currentText()
            
            rows = database.get_all_plan_search_items_for_export(seq, item, order, unit)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "计划检索数据"
            
            # Header
            ws.append(self.columns)
            
            # Data
            for row in rows:
                ws.append(list(row))
                
            wb.save(path)
            QMessageBox.information(self, "导出成功", f"已导出 {len(rows)} 条数据到\n{path}")
            
        except Exception as e:
            QMessageBox.critical(self, "导出出错", str(e))

    def on_sync_system(self):
        reply = QMessageBox.question(
            self, 
            "确认同步", 
            "将从系统现有订单明细中同步数据到检索库。\n仅同步序号不存在的记录或更新已有记录，不会删除额外数据。\n是否继续？",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                count = database.sync_plan_search_items_from_orders()
                QMessageBox.information(self, "同步成功", f"已处理 {count} 条系统记录")
                self.on_search()
            except Exception as e:
                QMessageBox.critical(self, "同步出错", str(e))
