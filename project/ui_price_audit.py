from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, 
    QTableWidget, QTableWidgetItem, QHeaderView, QTabWidget,
    QMessageBox, QDialog, QFileDialog, QAbstractItemView, QLineEdit,
    QInputDialog, QFormLayout, QStackedWidget
)
from PySide6.QtCore import Qt
from datetime import datetime
import database
import openpyxl
import price_analysis
from ui_plan_export import PlanExportWidget, MoneyDelegate

class PriceAuditModule(QWidget):
    def __init__(self):
        super().__init__()
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)
        
        # Tab 1: Plan Export (Existing Functionality)
        self.plan_export = PlanExportWidget()
        self.tabs.addTab(self.plan_export, "计划导出")
        
        # Tab 2: Other Quote Audit (New Functionality)
        self.other_audit = OtherQuoteAuditWidget()
        self.tabs.addTab(self.other_audit, "其他报价审核")


class OtherQuoteAuditWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        self.stack = QStackedWidget()
        layout.addWidget(self.stack)

        # Page 0: List
        self.page_list = QWidget()
        self.setup_list_ui(self.page_list)
        self.stack.addWidget(self.page_list)

        # Page 1: Detail
        self.page_detail = QuoteAuditDetailWidget(self)
        self.stack.addWidget(self.page_detail)

    def setup_list_ui(self, parent):
        layout = QVBoxLayout(parent)
        
        # Toolbar
        toolbar = QHBoxLayout()
        btn_add = QPushButton("新增审核记录")
        btn_add.clicked.connect(self.add_record)
        toolbar.addWidget(btn_add)
        
        btn_refresh = QPushButton("刷新")
        btn_refresh.clicked.connect(self.load_data)
        toolbar.addWidget(btn_refresh)
        
        btn_del = QPushButton("删除选中")
        btn_del.clicked.connect(self.delete_record)
        toolbar.addWidget(btn_del)
        
        toolbar.addStretch()
        layout.addLayout(toolbar)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["ID", "审核名称", "创建时间", "状态", "备注"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.cellDoubleClicked.connect(self.open_detail)
        layout.addWidget(self.table)
        
    def load_data(self):
        self.table.setRowCount(0)
        records = database.get_quote_audit_records()
        self.table.setRowCount(len(records))
        for r, row in enumerate(records):
            # id, name, created_at, status, remark
            self.table.setItem(r, 0, QTableWidgetItem(str(row[0])))
            self.table.setItem(r, 1, QTableWidgetItem(str(row[1])))
            self.table.setItem(r, 2, QTableWidgetItem(str(row[2])))
            self.table.setItem(r, 3, QTableWidgetItem(str(row[3])))
            self.table.setItem(r, 4, QTableWidgetItem(str(row[4])))

    def add_record(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("新增审核记录")
        layout = QVBoxLayout(dlg)
        form = QFormLayout()
        name_input = QLineEdit()
        remark_input = QLineEdit()
        form.addRow("审核名称:", name_input)
        form.addRow("备注:", remark_input)
        layout.addLayout(form)
        
        btns = QHBoxLayout()
        btn_ok = QPushButton("确定")
        btn_ok.clicked.connect(dlg.accept)
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(dlg.reject)
        btns.addWidget(btn_ok)
        btns.addWidget(btn_cancel)
        layout.addLayout(btns)
        
        if dlg.exec():
            name = name_input.text().strip()
            remark = remark_input.text().strip()
            if not name:
                QMessageBox.warning(self, "提示", "请输入名称")
                return
            
            if database.create_quote_audit_record(name, remark=remark):
                self.load_data()
            else:
                QMessageBox.warning(self, "错误", "创建失败")

    def delete_record(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一行")
            return
            
        rid = self.table.item(row, 0).text()
        if QMessageBox.question(self, "确认", "确定删除选中记录及其明细吗？") == QMessageBox.Yes:
            if database.delete_quote_audit_record(rid):
                self.load_data()
            else:
                QMessageBox.warning(self, "错误", "删除失败")

    def open_detail(self, row, col):
        rid = self.table.item(row, 0).text()
        name = self.table.item(row, 1).text()
        self.page_detail.load_record(rid, name)
        self.stack.setCurrentIndex(1)

    def go_back(self):
        self.stack.setCurrentIndex(0)
        self.load_data()


class QuoteAuditDetailWidget(QWidget):
    def __init__(self, parent_controller=None):
        super().__init__()
        self.parent_controller = parent_controller
        self.record_id = None
        self.record_name = ""
        self.setup_ui()
        
    def load_record(self, record_id, record_name):
        self.record_id = int(record_id)
        self.record_name = record_name
        self.lbl_title.setText(f"审核明细 - {record_name}")
        self.load_data()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Header
        header = QHBoxLayout()
        btn_back = QPushButton("返回列表")
        btn_back.clicked.connect(self.on_back)
        header.addWidget(btn_back)
        
        self.lbl_title = QLabel("审核明细")
        self.lbl_title.setStyleSheet("font-size: 16px; font-weight: bold; margin-left: 10px;")
        header.addWidget(self.lbl_title)
        header.addStretch()
        layout.addLayout(header)
        
        # Toolbar
        toolbar = QHBoxLayout()
        btn_import = QPushButton("导入Excel")
        btn_import.clicked.connect(self.import_excel)
        toolbar.addWidget(btn_import)
        
        btn_save = QPushButton("保存修改")
        btn_save.clicked.connect(self.save_data)
        toolbar.addWidget(btn_save)
        
        btn_smart = QPushButton("智能填充审核价")
        btn_smart.clicked.connect(self.smart_fill_prices)
        btn_smart.setStyleSheet("background-color: #e0f2f1; color: #00695c; font-weight: bold;")
        toolbar.addWidget(btn_smart)
        
        btn_complete = QPushButton("审核完成")
        btn_complete.clicked.connect(self.complete_audit)
        toolbar.addWidget(btn_complete)
        
        btn_export = QPushButton("导出Excel")
        btn_export.clicked.connect(self.export_excel)
        toolbar.addWidget(btn_export)
        
        toolbar.addStretch()
        layout.addLayout(toolbar)
        
        # Table
        self.table = QTableWidget()
        self.columns = [
            "序号", "主单编号", "需求单位", "采购标的", "规格型号",
            "单位", "采购数量", "预算(万)",
            "采购方式", "采购渠道", "计划发放", "询价金额", "审核金额", "备注"
        ]
        self.table.setColumnCount(len(self.columns))
        self.table.setHorizontalHeaderLabels(self.columns)
        self.table.setItemDelegateForColumn(11, MoneyDelegate(self.table))
        self.table.setItemDelegateForColumn(12, MoneyDelegate(self.table))
        
        # Set widths
        self.table.setColumnWidth(0, 80)
        self.table.setColumnWidth(1, 120)
        self.table.setColumnWidth(3, 150)
        self.table.setColumnWidth(4, 150)
        
        layout.addWidget(self.table)
        
    def on_back(self):
        if self.parent_controller:
            self.parent_controller.go_back()

    def smart_fill_prices(self):
        if self.table.rowCount() == 0:
            return

        from PySide6.QtGui import QCursor
        self.setCursor(QCursor(Qt.WaitCursor))
        
        count = 0
        errors = []
        
        try:
            for r in range(self.table.rowCount()):
                # Check if audit price (col 12) is empty
                item_audit = self.table.item(r, 12)
                current_val = item_audit.text().strip() if item_audit else ""
                
                # If has value > 0, skip
                try:
                    if current_val and float(current_val.replace(",", "")) > 0:
                        continue
                except:
                    pass # Empty or invalid, proceed
                
                # Get Item Name (col 3) and Spec (col 4)
                item_name_widget = self.table.item(r, 3)
                spec_model_widget = self.table.item(r, 4)
                item_name = item_name_widget.text().strip() if item_name_widget else ""
                spec_model = spec_model_widget.text().strip() if spec_model_widget else ""
                
                try:
                    rec = price_analysis.get_recommendation(item_name, spec_model)
                    if rec:
                        unit_price = rec.get('price', 0.0)
                        
                        # Get Qty (col 6)
                        qty_widget = self.table.item(r, 6)
                        qty_str = qty_widget.text().replace(",", "").strip() if qty_widget else "0"
                        try:
                            qty = float(qty_str)
                        except:
                            qty = 0.0
                            
                        total_price = unit_price * qty
                        
                        # Update UI
                        price_str = f"{total_price:.2f}"
                        if self.table.item(r, 12):
                            self.table.item(r, 12).setText(price_str)
                        else:
                            self.table.setItem(r, 12, QTableWidgetItem(price_str))
                        count += 1
                except Exception as e:
                    errors.append(f"Row {r+1}: {e}")
                    
        except Exception as e:
            errors.append(f"General error: {e}")
            
        self.setCursor(QCursor(Qt.ArrowCursor))
        
        if count > 0:
            QMessageBox.information(self, "完成", f"已智能填充 {count} 条数据的审核金额。\n请记得点击'保存修改'。")
        else:
            if not errors:
                QMessageBox.information(self, "提示", "未找到可填充的推荐价格。")
            else:
                QMessageBox.warning(self, "警告", f"发生错误:\n" + "\n".join(errors[:5]))
        
    def load_data(self):
        self.table.setRowCount(0)
        details = database.get_quote_audit_details(self.record_id)
        self.table.setRowCount(len(details))
        
        for r, row in enumerate(details):
            # row: id, detail_no, order_number, demand_unit, item_name, spec_model, 
            # unit, qty, budget, purchase_method, purchase_channel, plan_release, 
            # inquiry_price, audit_price, remark
            
            # Map to columns
            vals = [
                row[1], row[2], row[3], row[4], row[5],
                row[6], row[7], row[8], row[9], row[10], row[11], row[12], row[13], row[14]
            ]
            
            for c, v in enumerate(vals):
                txt = str(v) if v is not None else ""
                item = QTableWidgetItem(txt)
                
                if c in [11, 12]: # Editable
                    item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                else:
                    item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                
                self.table.setItem(r, c, item)
            
            # Store ID in first item
            self.table.item(r, 0).setData(Qt.UserRole, row[0])

    def import_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel Files (*.xlsx *.xls)")
        if not file_path:
            return
            
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Assuming header is row 1, data starts row 2
            # Try to map columns by name if possible, otherwise assume order
            # The user said "same fields as Plan Export module"
            # Let's try to find headers
            
            headers = {}
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                if val:
                    headers[str(val).strip()] = col
            
            # Mapping
            # "序号", "主单编号", "需求单位", "采购标的", "规格型号",
            # "单位", "采购数量", "预算(万)",
            # "采购方式", "采购渠道", "计划发放", "询价金额", "审核金额", "备注"
            
            map_cols = {
                "序号": "detail_no",
                "主单编号": "order_number",
                "需求单位": "demand_unit",
                "采购标的": "item_name",
                "规格型号": "spec_model",
                "单位": "unit",
                "采购数量": "qty",
                "预算(万)": "budget",
                "采购方式": "purchase_method",
                "采购渠道": "purchase_channel",
                "计划发放": "plan_release",
                "询价金额": "inquiry_price",
                "审核金额": "audit_price",
                "备注": "remark"
            }
            
            data_to_add = []
            
            for r in range(2, ws.max_row + 1):
                # Check if empty row
                if not ws.cell(row=r, column=1).value and not ws.cell(row=r, column=4).value:
                    continue
                    
                row_data = {}
                for header_name, key in map_cols.items():
                    col_idx = headers.get(header_name)
                    if col_idx:
                        val = ws.cell(row=r, column=col_idx).value
                        if val is None: val = ""
                        row_data[key] = val
                    else:
                        # Fallback: if headers not found, try to use index based on standard columns order?
                        # It's risky. Let's rely on headers for now.
                        # If headers missing, maybe we should try standard indices
                        pass
                
                # If map_cols didn't match anything, maybe use indices
                if not row_data and len(headers) < 3: 
                    # Try indices 1..14
                    vals = []
                    for c in range(1, 15):
                        vals.append(ws.cell(row=r, column=c).value)
                    
                    row_data = {
                        "detail_no": vals[0], "order_number": vals[1], "demand_unit": vals[2],
                        "item_name": vals[3], "spec_model": vals[4], "unit": vals[5],
                        "qty": vals[6], "budget": vals[7], "purchase_method": vals[8],
                        "purchase_channel": vals[9], "plan_release": vals[10],
                        "inquiry_price": vals[11], "audit_price": vals[12], "remark": vals[13]
                    }

                if row_data:
                    data_to_add.append(row_data)
            
            if data_to_add:
                if database.add_quote_audit_details(self.record_id, data_to_add):
                    QMessageBox.information(self, "成功", f"导入 {len(data_to_add)} 条数据")
                    self.load_data()
                else:
                    QMessageBox.warning(self, "失败", "保存数据失败")
            else:
                QMessageBox.warning(self, "提示", "未读取到有效数据，请检查Excel表头是否与系统一致")
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导入出错: {str(e)}")

    def save_data(self):
        # Save modified audit prices
        count = 0
        try:
            for r in range(self.table.rowCount()):
                did = self.table.item(r, 0).data(Qt.UserRole)
                audit_price_item = self.table.item(r, 12)
                if audit_price_item:
                    txt = audit_price_item.text().replace(",", "")
                    try:
                        val = float(txt)
                    except:
                        val = 0.0
                    
                    # Always update for now, or check dirty?
                    # Ideally check dirty, but simple update is fine
                    database.update_quote_audit_detail_price(did, val)
                    count += 1
            QMessageBox.information(self, "成功", f"保存成功")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"保存失败: {e}")

    def complete_audit(self):
        if QMessageBox.question(self, "确认", "确定标记为审核完成吗？") == QMessageBox.Yes:
            database.update_quote_audit_status(self.record_id, "审核完成")
            QMessageBox.information(self, "成功", "状态已更新")
            self.on_back()

    def export_excel(self):
        if self.table.rowCount() == 0:
            return
            
        default_name = f"{self.record_name}_审核明细.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(self, "导出Excel", default_name, "Excel Files (*.xlsx)")
        
        if not file_path:
            return
            
        # Collect data
        rows = []
        for r in range(self.table.rowCount()):
            row_data = []
            for c in range(self.table.columnCount()):
                item = self.table.item(r, c)
                row_data.append(item.text() if item else "")
            rows.append(row_data)
            
        # Header info
        header_info = {
            "number": self.record_name,
            "task_name": "报价审核",
            "unit": "多部门",
            "yymm": datetime.now().strftime("%Y-%m"),
            "purchaser": "审核员"
        }
        
        try:
            from export import OrderExporter
            exporter = OrderExporter(header_info, self.columns, rows, title=f"审核明细: {self.record_name}")
            exporter.export(file_path)
            QMessageBox.information(self, "成功", f"导出成功:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
