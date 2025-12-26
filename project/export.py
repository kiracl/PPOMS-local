import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

class OrderExporter:
    def __init__(self, header_info: dict, columns: list, rows: list, title: str = "采购计划表"):
        # ... existing code ...
        self.header_info = header_info
        self.columns = columns
        self.rows = rows
        self.title = title

    def export(self, filepath: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "采购计划"
        
        # --- Styles ---
        border_thin = Side(border_style="thin", color="000000")
        border_all = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)
        
        font_title = Font(name="SimSun", size=18, bold=True)
        font_header = Font(name="SimSun", size=11)
        font_table_header = Font(name="SimSun", size=11, bold=True)
        font_cell = Font(name="SimSun", size=10)
        font_group_header = Font(name="SimSun", size=11, bold=True)
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        fill_semi = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid") # RGB 240,240,240
        fill_civil = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # RGB 220,230,241
        
        # --- 1. Title ---
        total_cols = len(self.columns)
        if total_cols < 1: total_cols = 1
        
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        cell = ws.cell(row=1, column=1, value=self.title)
        cell.font = font_title
        cell.alignment = align_center
        ws.row_dimensions[1].height = 40
        
        # --- 2. Header Info ---
        # (Assuming we skip detailed header info for export plan as per previous request context if needed, 
        # but here we keep generic logic or adapt based on header_info content)
        # Detect export plan by first column name
        is_export_plan = (len(self.columns) > 0 and self.columns[0] == "序号")
        
        row_idx = 2
        if not is_export_plan:
             # ... existing header logic ...
             # Row 2
            info_str_1 = f"主单编号：{self.header_info.get('number', '')}"
            info_str_2 = f"采购任务名称：{self.header_info.get('task_name', '')}"
            
            mid_point = total_cols // 2
            
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mid_point)
            c = ws.cell(row=2, column=1, value=info_str_1)
            c.font = font_header
            c.alignment = align_left
            
            ws.merge_cells(start_row=2, start_column=mid_point+1, end_row=2, end_column=total_cols)
            c = ws.cell(row=2, column=mid_point+1, value=info_str_2)
            c.font = font_header
            c.alignment = align_left
            
            # Row 3
            info_str_3 = f"需求单位：{self.header_info.get('unit', '')}"
            info_str_4 = f"计划月份：{self.header_info.get('yymm', '')}"
            info_str_5 = f"采购员：{self.header_info.get('purchaser', '')}"
            
            col_span = total_cols // 3
            if col_span < 1: col_span = 1
            
            # Unit
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_span)
            c = ws.cell(row=3, column=1, value=info_str_3)
            c.font = font_header
            c.alignment = align_left
            
            # Month
            ws.merge_cells(start_row=3, start_column=col_span+1, end_row=3, end_column=col_span*2)
            c = ws.cell(row=3, column=col_span+1, value=info_str_4)
            c.font = font_header
            c.alignment = align_left
            
            # Purchaser
            ws.merge_cells(start_row=3, start_column=col_span*2+1, end_row=3, end_column=total_cols)
            c = ws.cell(row=3, column=col_span*2+1, value=info_str_5)
            c.font = font_header
            c.alignment = align_left
            
            ws.row_dimensions[2].height = 25
            ws.row_dimensions[3].height = 25
            row_idx = 4
        else:
            # For Export Plan, skip header info rows
            row_idx = 2

        # --- 3. Table Header ---
        for col_idx, col_name in enumerate(self.columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=col_name)
            cell.font = font_table_header
            cell.border = border_all
            cell.alignment = align_center
            
            # Heuristic column width
            width = 12
            if "序号" in col_name: width = 8
            elif "名称" in col_name or "任务" in col_name or "备注" in col_name or "规格" in col_name: width = 25
            elif "数量" in col_name or "单价" in col_name or "金额" in col_name: width = 12
            ws.column_dimensions[get_column_letter(col_idx)].width = width
                
        ws.row_dimensions[row_idx].height = 30
        
        # --- 4. Table Rows with Grouping ---
        row_idx += 1
        
        # Flags to track if we have inserted the group header
        inserted_semi_header = False
        inserted_civil_header = False
        
        for row_data in self.rows:
            # Determine group based on detail number (1st column)
            # Format: 2601MPB-1 (Semi), 2601MP-1 (Civil), 2601MPJ-1 (Machined)
            detail_no = str(row_data[0]) if len(row_data) > 0 else ""
            
            is_semi = "MPB" in detail_no
            is_civil = "MP-" in detail_no or (detail_no.endswith("MP") if "MP" in detail_no else False) or ("MP" in detail_no and "MPB" not in detail_no and "MPJ" not in detail_no)
            
            # Insert Semi Header
            if is_semi and not inserted_semi_header:
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
                c = ws.cell(row=row_idx, column=1, value="半成品MPB")
                c.font = font_group_header
                c.alignment = align_left  # Left align as requested
                c.fill = fill_semi
                c.border = border_all
                # Apply border to all merged cells
                for i in range(2, total_cols + 1):
                    c2 = ws.cell(row=row_idx, column=i)
                    c2.border = border_all
                    c2.fill = fill_semi
                
                ws.row_dimensions[row_idx].height = 25
                row_idx += 1
                inserted_semi_header = True
                
            # Insert Civil Header
            if is_civil and not inserted_civil_header:
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
                c = ws.cell(row=row_idx, column=1, value="民品MP")
                c.font = font_group_header
                c.alignment = align_left  # Left align as requested
                c.fill = fill_civil
                c.border = border_all
                for i in range(2, total_cols + 1):
                    c2 = ws.cell(row=row_idx, column=i)
                    c2.border = border_all
                    c2.fill = fill_civil
                    
                ws.row_dimensions[row_idx].height = 25
                row_idx += 1
                inserted_civil_header = True
            
            # Write Row Data
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = font_cell
                cell.border = border_all
                cell.alignment = align_center
            row_idx += 1
            
        # --- 5. Footer ---
        row_idx += 1
        footer_1 = "审核："
        footer_2 = "签收："
        footer_3 = "编制："
        
        col_span = total_cols // 3
        if col_span < 1: col_span = 1
        
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=col_span)
        c = ws.cell(row=row_idx, column=1, value=footer_1)
        c.font = font_header
        c.alignment = align_left
        
        ws.merge_cells(start_row=row_idx, start_column=col_span+1, end_row=row_idx, end_column=col_span*2)
        c = ws.cell(row=row_idx, column=col_span+1, value=footer_2)
        c.font = font_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells(start_row=row_idx, start_column=col_span*2+1, end_row=row_idx, end_column=total_cols)
        c = ws.cell(row=row_idx, column=col_span*2+1, value=footer_3)
        c.font = font_header
        c.alignment = Alignment(horizontal="right", vertical="center")
        
        ws.row_dimensions[row_idx].height = 30
        
        wb.save(filepath)

def write_detail_import_template(path: str, purchasers_hint: bool = False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "导入模板"
    headers = [
        "采购标的",
        "规格型号",
        "采购数量",
        "单位",
        "单价(元)",
        "采购方式",
        "采购途径",
        "计划发放",
        "备注",
    ]
    ws.append(headers)
    ws.append([
        "示例：激光测距仪",
        "Leica-D510",
        "1",
        "个",
        "100.00",
        "框架协议",
        "能建商城",
        "未分配",
        "示例备注：支持中英文与换行\n第二行说明",
    ])
    from openpyxl.styles import Alignment
    for idx, name in enumerate(headers, start=1):
        width = 14
        if name in ("采购标的", "规格型号", "备注"):
            width = 28
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws["I2"].alignment = Alignment(wrap_text=True)

    info = wb.create_sheet("说明")
    lines = [
        "数据导入模板使用说明",
        "",
        "必填列：采购标的、规格型号、采购数量、单位、单价(元)、采购方式、采购途径、计划发放",
        "备注列：可选填写，支持中英文、常用符号及换行；导入时长度限制为500字符，超过部分将自动截断并在导入提示中告知",
        "采购方式选项：询比采购/公开招标/集中采购/框架协议",
        "采购途径选项：能建商城/采购平台/线下采购",
        "计划发放：留空或填‘未分配’表示未分配，系统可根据采购标的自动推荐",
    ]
    for r, text in enumerate(lines, start=1):
        info.cell(row=r, column=1, value=text)
    wb.save(path)
